#!/usr/bin/env python3
"""Regenerate src/employsi/data/usVacancyDemand.ts from two US Bureau of Labor
Statistics releases, pulled live from BLS's own public distribution:

  • OEWS  — Occupational Employment and Wage Statistics, metropolitan-area file
            (https://www.bls.gov/oes/special-requests/oesmYYma.zip). Gives
            EMPLOYMENT by detailed SOC occupation × MSA, annually (May reference).
            This supplies the occupational and geographic STRUCTURE.
  • JOLTS — Job Openings and Labor Turnover Survey, via the BLS Public Data API
            (https://api.bls.gov/publicAPI/v1/timeseries/data/). Gives the
            seasonally-adjusted monthly JOB OPENINGS RATE per state. This
            supplies the LEVEL and the month-to-month path.

Why both: OEWS is an employment stock with no vacancy measure, and JOLTS has a
real vacancy measure but no occupational detail. Neither alone is comparable
with the app's other feeds (AU IVI, UK ONS, EU Eurostat), which are vacancy
COUNTS by occupation by place. Combining them gives exactly that:

    vacancies(skill, metro, month)
        = employment(skill, metro, year-of-month) × r / (1 − r)

where r is that metro's STATE job-openings rate for that month. The r/(1−r) term
is not a fudge: JOLTS defines the openings rate as openings ÷ (employment +
openings), so inverting it recovers openings from an employment base. The one
modelling assumption is that an occupation's vacancy rate matches its state's
all-industry rate — the same assumption the EU wiring makes with its flat 2.5%,
except here the rate is real, monthly, and state-specific rather than a constant.

Each detailed SOC occupation title is term-matched to canonical skills through
the shared skillsTaxonomy terms and summed per skill, exactly as the UK and EU
wirings do, so the US counts land on the same skill vocabulary as everything
else. Employment is recorded per city (metro), not rolled up to the country —
the global layer's country roll-up does that itself from COUNTRY_MEMBERS.

Coverage: OEWS vintages 2019..latest. Before 2019 the biggest metros (New York,
Chicago, Los Angeles, Boston, Dallas, Washington, Philadelphia, Seattle, San
Francisco) are published only as metropolitan DIVISIONS or NECTAs rather than
whole MSAs, so a 2014-2018 series would silently compare different footprints
year to year. Those months are left at zero rather than back-filled.

Usage: python3 scripts/gen-us-vacancy-demand.py [--cache DIR] [--from 2019]
Network required (BLS downloads + API). No API key needed: the v1 API is
unauthenticated, and we stay well inside its 25-series / 10-year / 25-queries
per-day limits (17 state series, three decade windows).
"""
from __future__ import annotations
import glob
import json
import os
import re
import sys
import urllib.request
import zipfile

sys.path.insert(0, __file__.rsplit('/', 1)[0])
from skills_taxonomy import matcher  # noqa: E402

ROOT = __file__.rsplit('/scripts/', 1)[0]
TAX = f'{ROOT}/src/employsi/data/skillsTaxonomy.ts'
MONTHS_TS = f'{ROOT}/src/employsi/data/iviSkillDemand.ts'
OUT = f'{ROOT}/src/employsi/data/usVacancyDemand.ts'

# BLS asks for a descriptive User-Agent with contact details on bulk downloads;
# requests without one are refused.
UA = 'employsi-datafetch/1.0 (labour-market map; contact via repo)'
OEWS_URL = 'https://www.bls.gov/oes/special-requests/oesm{yy}ma.zip'
JOLTS_API = 'https://api.bls.gov/publicAPI/v1/timeseries/data/'

# ── City → BLS geography ─────────────────────────────────────────────────────
# msa:   OEWS area codes accepted for this city, in preference order. Most are a
#        single stable CBSA; Boston needs two because OEWS published it as the
#        Boston-Cambridge-Nashua NECTA (71650) up to 2023 and as the
#        Boston-Cambridge-Newton MSA (14460) from 2024. Only the FIRST code
#        present in a given vintage is used, so the two never double-count.
# state: FIPS code for the JOLTS state series supplying the monthly openings rate.
CITIES: dict[str, dict] = {
    'newyork':      {'msa': ['35620'], 'state': '36'},  # NY
    'losangeles':   {'msa': ['31080'], 'state': '06'},  # CA
    'chicago':      {'msa': ['16980'], 'state': '17'},  # IL
    'dallas':       {'msa': ['19100'], 'state': '48'},  # TX
    'houston':      {'msa': ['26420'], 'state': '48'},  # TX
    'washington':   {'msa': ['47900'], 'state': '11'},  # DC
    'philadelphia': {'msa': ['37980'], 'state': '42'},  # PA
    'atlanta':      {'msa': ['12060'], 'state': '13'},  # GA
    'boston':       {'msa': ['14460', '71650'], 'state': '25'},  # MA
    'sanfrancisco': {'msa': ['41860'], 'state': '06'},  # CA
    # San Jose is its own MSA, NOT part of San Francisco's — which is exactly why
    # the Silicon Valley companies had to be moved off the San Francisco pin.
    'sanjose':      {'msa': ['41940'], 'state': '06'},  # CA
    'seattle':      {'msa': ['42660'], 'state': '53'},  # WA
    'minneapolis':  {'msa': ['33460'], 'state': '27'},  # MN
    'sandiego':     {'msa': ['41740'], 'state': '06'},  # CA
    'denver':       {'msa': ['19740'], 'state': '08'},  # CO
    'charlotte':    {'msa': ['16740'], 'state': '37'},  # NC
    'austin':       {'msa': ['12420'], 'state': '48'},  # TX
    'portland':     {'msa': ['38900'], 'state': '41'},  # OR
    'cincinnati':   {'msa': ['17140'], 'state': '39'},  # OH
    'indianapolis': {'msa': ['26900'], 'state': '18'},  # IN
    'omaha':        {'msa': ['36540'], 'state': '31'},  # NE
    # Bentonville (Walmart's home town) is not its own MSA — it sits inside the
    # Fayetteville-Springdale-Rogers, AR metro, which is the published area.
    'bentonville':  {'msa': ['22220'], 'state': '05'},  # AR
}

STATE_NAME = {
    '05': 'Arkansas', '06': 'California', '08': 'Colorado',
    '11': 'District of Columbia', '13': 'Georgia', '17': 'Illinois',
    '18': 'Indiana', '25': 'Massachusetts', '27': 'Minnesota',
    '31': 'Nebraska', '36': 'New York', '37': 'North Carolina',
    '39': 'Ohio', '41': 'Oregon', '42': 'Pennsylvania', '48': 'Texas',
    '53': 'Washington',
}

# OEWS aggregation levels: only 'detailed' rows are disjoint, so summing them is
# safe. 'total'/'major'/'minor'/'broad' are roll-ups of the same jobs.
DETAILED = 'detailed'


def http(url: str, timeout: int = 120) -> bytes:
    req = urllib.request.Request(url, headers={'User-Agent': UA})
    return urllib.request.urlopen(req, timeout=timeout).read()


def load_months() -> list[str]:
    """The shared IVI_MONTHS axis every country series is aligned to."""
    txt = open(MONTHS_TS).read()
    m = re.search(r'export const IVI_MONTHS[^=]*=\s*(\[.*?\]);', txt, re.S)
    if not m:
        sys.exit('Could not read IVI_MONTHS from iviSkillDemand.ts')
    return json.loads(m.group(1).replace("'", '"'))


# ── OEWS ─────────────────────────────────────────────────────────────────────
def oews_path(cache: str, year: int) -> str | None:
    """Download (once) and unpack the OEWS metro file for `year`, returning the
    path to its MSA workbook, or None when that vintage isn't published yet."""
    yy = f'{year % 100:02d}'
    zpath = f'{cache}/oesm{yy}ma.zip'
    if not os.path.exists(zpath):
        url = OEWS_URL.format(yy=yy)
        try:
            data = http(url, timeout=600)
        except Exception as e:  # noqa: BLE001
            sys.stderr.write(f'  {year}: not available ({e})\n')
            return None
        # A vintage that isn't published yet answers 200 with an HTML error page
        # rather than a 404, so check the payload really is an archive before
        # caching it — otherwise the bad file poisons every later run.
        if not data.startswith(b'PK'):
            sys.stderr.write(f'  {year}: not published yet (server returned '
                             f'{len(data)} bytes of non-archive content)\n')
            return None
        open(zpath, 'wb').write(data)
    dest = f'{cache}/oes{yy}'
    if not os.path.isdir(dest):
        with zipfile.ZipFile(zpath) as z:
            z.extractall(dest)
    hits = glob.glob(f'{dest}/**/MSA_M{year}_dl.xlsx', recursive=True)
    if not hits:
        sys.stderr.write(f'  {year}: no single MSA workbook in the archive (pre-2014 layout)\n')
        return None
    return hits[0]


def read_oews(path: str, skills_for, want_codes: set[str]):
    """Return (emp[skill][area_code], areas_present) for the OEWS area codes in
    `want_codes`. Keyed by AREA rather than by city because a city can list
    several accepted codes (Boston's NECTA vs MSA) — the caller picks exactly
    one per vintage so the two can never be summed together."""
    import openpyxl  # imported late so --help works without the dependency

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    hdr = [str(h or '').strip().upper() for h in next(rows)]

    def col(*names):
        for n in names:
            if n in hdr:
                return hdr.index(n)
        sys.exit(f'OEWS column not found: {names} in {hdr[:12]}')

    c_area = col('AREA')
    c_occ = col('OCC_TITLE')
    c_grp = col('O_GROUP', 'OCC_GROUP')
    c_emp = col('TOT_EMP')

    emp: dict[str, dict[str, int]] = {}
    areas: set[str] = set()
    for r in rows:
        area = str(r[c_area] or '').strip()
        areas.add(area)
        if area not in want_codes or str(r[c_grp] or '').strip().lower() != DETAILED:
            continue
        v = r[c_emp]
        if not isinstance(v, (int, float)):
            continue  # '**' = suppressed for disclosure / reliability
        for s in skills_for(str(r[c_occ] or '')):
            emp.setdefault(s, {}).setdefault(area, 0)
            emp[s][area] += int(v)
    return emp, areas


# ── JOLTS ────────────────────────────────────────────────────────────────────
def jolts_series_id(state: str) -> str:
    """JT | S(seasonally adjusted) | industry(6) | state(2) | area(5) | size(2)
    | element(2) | ratelevel(1) — total nonfarm job-openings RATE for a state."""
    return f'JTS000000{state}0000000JOR'


def fetch_jolts(states: list[str], first_year: int, last_year: int):
    """{state: {'YYYY-MM': rate_pct}} from the unauthenticated v1 API, which
    caps each request at 25 series and 10 years."""
    out: dict[str, dict[str, float]] = {s: {} for s in states}
    ids = [jolts_series_id(s) for s in states]
    by_id = {jolts_series_id(s): s for s in states}
    y = first_year
    while y <= last_year:
        end = min(y + 9, last_year)
        body = json.dumps({'seriesid': ids, 'startyear': str(y), 'endyear': str(end)}).encode()
        req = urllib.request.Request(
            JOLTS_API, data=body,
            headers={'Content-type': 'application/json', 'User-Agent': UA},
        )
        j = json.loads(urllib.request.urlopen(req, timeout=120).read())
        if j.get('status') != 'REQUEST_SUCCEEDED':
            sys.exit(f'BLS API refused {y}-{end}: {j.get("status")} {j.get("message")}')
        for s in j['Results']['series']:
            st = by_id.get(s['seriesID'])
            if not st:
                continue
            for d in s['data']:
                p = d['period']
                if not p.startswith('M') or p == 'M13':
                    continue  # M13 is the annual average
                try:
                    out[st][f'{d["year"]}-{p[1:]}'] = float(d['value'])
                except ValueError:
                    pass
        sys.stderr.write(f'  JOLTS {y}-{end}: '
                         f'{sum(len(v) for v in out.values())} state-months so far\n')
        y = end + 1
    return out


def rate_array(by_month: dict[str, float], months: list[str], first_year: int) -> list[float]:
    """Openings rate (as a percentage) on the shared month axis: zero before the
    first OEWS year, forward-filled through any gap and past the last published
    month so the series runs to the end of the axis."""
    out: list[float] = []
    last = 0.0
    for ym in months:
        if int(ym[:4]) < first_year:
            out.append(0.0)
            continue
        v = by_month.get(ym)
        if v is not None:
            last = v
        out.append(last)
    return out


def main() -> int:
    args = sys.argv[1:]
    cache = args[args.index('--cache') + 1] if '--cache' in args else '/tmp/bls-cache'
    first_year = int(args[args.index('--from') + 1]) if '--from' in args else 2019
    os.makedirs(cache, exist_ok=True)

    # One shared, validated matcher — the same one the AU/UK/EU generators use.
    skills_for = matcher(TAX)
    months = load_months()
    axis_last_year = int(months[-1][:4])

    want_codes = {c for cfg in CITIES.values() for c in cfg['msa']}

    sys.stderr.write(f'OEWS metro files ({first_year}..):\n')
    per_year: dict[int, dict[str, dict[str, int]]] = {}
    for year in range(first_year, axis_last_year + 1):
        path = oews_path(cache, year)
        if not path:
            continue
        by_area, areas = read_oews(path, skills_for, want_codes)

        # A city may accept several area codes across vintages (Boston). Take
        # only the FIRST one present in THIS vintage, so the alternatives can
        # never be summed together.
        chosen: dict[str, str] = {}
        for cid, cfg in CITIES.items():
            present = [c for c in cfg['msa'] if c in areas]
            if present:
                chosen[cid] = present[0]
            else:
                sys.stderr.write(f'  {year}: {cid} absent from this vintage — skipped\n')
        emp: dict[str, dict[str, int]] = {}
        for s, byArea in by_area.items():
            for cid, code in chosen.items():
                v = byArea.get(code, 0)
                if v:
                    emp.setdefault(s, {})[cid] = v
        per_year[year] = emp
        tot = sum(sum(v.values()) for v in emp.values())
        sys.stderr.write(f'  {year}: {len(emp)} skills, {len(chosen)} metros, '
                         f'{tot:,} matched jobs\n')

    if not per_year:
        sys.exit('No OEWS vintages could be read.')
    years = sorted(per_year)

    states = sorted({cfg['state'] for cfg in CITIES.values()})
    sys.stderr.write(f'JOLTS openings rate for {len(states)} states:\n')
    jolts = fetch_jolts(states, years[0], axis_last_year)
    missing = [s for s in states if not jolts.get(s)]
    if missing:
        sys.exit(f'No JOLTS data returned for states {missing} — refusing to '
                 f'substitute the national rate silently.')

    rates = {s: rate_array(jolts[s], months, years[0]) for s in states}

    # skill → city → [employment per year]
    order_key = years[-1]
    all_skills = sorted(
        {s for y in years for s in per_year[y]},
        key=lambda s: -sum(per_year[order_key].get(s, {}).values()),
    )

    L: list[str] = []
    A = L.append
    A('// GENERATED — do not edit by hand. Run scripts/gen-us-vacancy-demand.py.')
    A('// Source: US Bureau of Labor Statistics —')
    A('//   • OEWS (Occupational Employment and Wage Statistics), metropolitan-area')
    A('//     files, May {}..{}: employment by detailed SOC occupation × metro.'.format(years[0], years[-1]))
    A('//   • JOLTS (Job Openings and Labor Turnover Survey) via the BLS public API:')
    A('//     seasonally-adjusted monthly job-openings RATE by state.')
    A('//')
    A('// OEWS measures employment, not vacancies, so on its own it is not comparable')
    A("// with this app's other feeds (AU IVI, UK ONS, EU Eurostat), which are vacancy")
    A('// counts. JOLTS supplies the missing vacancy measure. Because JOLTS defines the')
    A('// openings rate r as openings ÷ (employment + openings), openings are recovered')
    A('// from an employment base as employment × r / (1 − r) — that is the identity,')
    A('// not an adjustment factor. Each metro uses its own state rate for the month in')
    A('// question, so the level and the monthly path are both real BLS data; the one')
    A("// assumption is that an occupation's vacancy rate matches its state's")
    A('// all-industry rate.')
    A('//')
    A('// Occupation titles are term-matched to canonical skills through the shared')
    A('// skillsTaxonomy terms, exactly as the UK and EU wirings do. Only OEWS')
    A("// 'detailed' rows are summed (the major/minor/broad rows re-count the same")
    A('// jobs), and suppressed cells are skipped rather than estimated.')
    A('//')
    A(f'// Months before {years[0]} are zero: earlier OEWS vintages publish the largest')
    A('// metros only as metropolitan divisions/NECTAs, so a longer series would be')
    A('// comparing different geographic footprints year to year.')
    A('')
    A("import { IVI_MONTHS } from './iviSkillDemand';")
    A('')
    A(f'export const US_YEARS: number[] = {json.dumps(years)};')
    A("export const US_SOURCE = 'US Bureau of Labor Statistics — OEWS employment by "
      "occupation and metro, scaled to vacancies by the JOLTS state job-openings rate';")
    A('')
    A('// City id → the state whose JOLTS openings rate scales it.')
    A('export const US_CITY_STATE: Record<string, string> = {')
    for cid, cfg in CITIES.items():
        A(f'  {cid}: {json.dumps(cfg["state"])}, // {STATE_NAME[cfg["state"]]}')
    A('};')
    A('')
    A('// State FIPS → monthly job-openings rate (%), aligned to IVI_MONTHS.')
    A('const US_STATE_RATE: Record<string, number[]> = {')
    for s in states:
        A(f'  {json.dumps(s)}: [{",".join(f"{v:g}" for v in rates[s])}],')
    A('};')
    A('')
    A('// Compact source values: skill → city → [OEWS employment per US_YEARS].')
    A('const US_EMPLOYMENT: Record<string, Record<string, number[]>> = {')
    for s in all_skills:
        parts = []
        for cid in CITIES:
            arr = [per_year[y].get(s, {}).get(cid, 0) for y in years]
            if any(arr):
                parts.append(f'{cid}: [{",".join(map(str, arr))}]')
        if parts:
            A(f'  {json.dumps(s)}: {{ {", ".join(parts)} }},')
    A('};')
    A('')
    A('// Expand onto the shared monthly axis: each month takes its calendar year\'s')
    A('// OEWS employment (the latest year carried forward to "now"), converted to')
    A('// active vacancies with that month\'s state openings rate.')
    A('// Month index → index into US_YEARS: the most recent published vintage at or')
    A('// before that month\'s calendar year, or -1 when the month predates the first')
    A('// vintage. Forward-filling this way means a gap year inherits the previous')
    A('// vintage and months after the last one carry it forward to "now".')
    A('const YEAR_SLOT: number[] = IVI_MONTHS.map((ym) => {')
    A('  const y = Number(ym.slice(0, 4));')
    A('  let slot = -1;')
    A('  for (let i = 0; i < US_YEARS.length; i++) if (US_YEARS[i] <= y) slot = i;')
    A('  return slot;')
    A('});')
    A('')
    A('function expand(byCity: Record<string, number[]>): Record<string, number[]> {')
    A('  const out: Record<string, number[]> = {};')
    A('  for (const [city, emp] of Object.entries(byCity)) {')
    A('    const rate = US_STATE_RATE[US_CITY_STATE[city]] || [];')
    A('    out[city] = YEAR_SLOT.map((yi, i) => {')
    A('      if (yi < 0) return 0;')
    A('      const e = emp[yi] || 0;')
    A('      const r = (rate[i] || 0) / 100;')
    A('      if (e <= 0 || r <= 0 || r >= 1) return 0;')
    A('      return Math.round((e * r) / (1 - r));')
    A('    });')
    A('  }')
    A('  return out;')
    A('}')
    A('')
    A('// Skill → metro → monthly vacancy history (aligned to IVI_MONTHS).')
    A('export const US_SERIES: Record<string, Record<string, number[]>> =')
    A('  Object.fromEntries(Object.entries(US_EMPLOYMENT).map(([s, e]) => [s, expand(e)]));')
    A('')
    A('// Skill → latest-month metro vacancy count (current heat map).')
    A('export const US_SKILL_BY_CITY: Record<string, Record<string, number>> =')
    A('  Object.fromEntries(')
    A('    Object.entries(US_SERIES).map(([s, byCity]) => [')
    A('      s,')
    A('      Object.fromEntries(Object.entries(byCity).map(([c, arr]) => [c, arr[arr.length - 1]])),')
    A('    ]),')
    A('  );')
    A('')
    open(OUT, 'w').write('\n'.join(L))

    # Report at the same scale the app will show, so the numbers can be sanity
    # checked against the UK/EU feeds without opening the app.
    last_rate = {s: rates[s][-1] for s in states}
    print(f'{len(all_skills)} skills × {len(CITIES)} metros, OEWS {years[0]}..{years[-1]} '
          f'-> {OUT}')
    print(f'latest state openings rates: '
          f'{", ".join(f"{STATE_NAME[s]} {last_rate[s]:g}%" for s in states[:4])}, …')
    top = []
    for s in all_skills[:8]:
        v = 0
        for cid, cfg in CITIES.items():
            e = per_year[years[-1]].get(s, {}).get(cid, 0)
            r = last_rate[cfg['state']] / 100
            v += round(e * r / (1 - r)) if 0 < r < 1 else 0
        top.append(f'{s} {v:,}')
    print(f'latest-month vacancies (all {len(CITIES)} metros): ' + '; '.join(top))
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
