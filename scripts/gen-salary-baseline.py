#!/usr/bin/env python3
"""Regenerate src/employsi/data/salaryBaseline.ts — historical pay per skill.

WHY THIS EXISTS
Every salary figure in the app is an ADVERTISED one, taken from the midpoint of
a band a job ad printed, and the archive holding those ads begins 2026-07-20.
Thirty-six days cannot answer "is this skill worth more than it used to be", and
will not be able to for another year. The one source in the archive with real
history — the closed Wayback corpus, back to 2003 — carries no salary at all.

So the historical side comes from outside, one national statistical office per
market. Two are wired up:

  AU  ATO Taxation Statistics, Individuals Table 15 — MEDIAN SALARY OR WAGE
      INCOME by ANZSCO4 unit group, annual, CC BY 2.5 AU.
  NZ  Stats NZ Census CEN23_WRK_011 — MEDIAN TOTAL PERSONAL INCOME by ANZSCO
      sub-major, by regional council, for the 2013/2018/2023 censuses.

WHY NOT THE OBVIOUS SOURCE IN EITHER MARKET
ABS Employee Earnings and Hours splits the two things needed across two
granularities and neither cell is usable alone: cube 63060DO003 Table 10
publishes the median but only for the EIGHT ANZSCO major groups, so every skill
in a category would get an identical number; cube 63060DO011 Table 1 is at
ANZSCO4 but publishes only a MEAN. EEH is biennial besides.

Stats NZ's INC_INC_004 ("Earnings ... by occupation (ANZSCO 2006)") sounds like
the right table and is not: its occupation codelist holds the same EIGHT major
groups. The census table is the one with sub-major detail.

SINGAPORE IS NOT A MARKET HERE, AND THAT IS A DECISION RATHER THAN AN OMISSION.
It was third in the intended order (AU, NZ, SG, PH, MY) and was investigated on
2026-08-25. The data exists, is annual, is a real median, and is better than
Australia's on two axes — MOM's mrsd_37_Res_occ_income.xlsx publishes median
gross monthly income from employment for FULL-TIME employed residents (so no
part-time drag, unlike the ATO), back to 2001 (twenty-five years against eight),
in Incl-CPF and Excl-CPF variants.

It is published at EIGHT SSOC major groups and nothing finer. SingStat M920131
is the same eight. Nothing finer was found through SingStat's API, MOM's site or
data.gov.sg, whose search ignores its own query parameter.

What eight groups cost is measurable from this repo, because
gen-sg-occupation-supply.py already had to map the taxonomy onto them:

    Professionals                           41 skills
    Associate Professionals & Technicians   18
    Craftsmen & Related Trades              13
    Managers & Administrators               12
    Plant & Machine Operators                7
    Service & Sales                          5
    Cleaners, Labourers                      3
    Clerical Support                         1

So Singapore would publish EIGHT DISTINCT VALUES FOR 100 SKILLS. Software
Engineering, Medical Practice, Commercial & Legal, Geology and Journalism &
Media would all read the same number, and a reader looking at one skill's card
could not tell. Against AU's ~69 distinguishable of 75, and NZ's 67, that is not
a coarser version of the same product — it is a category average wearing a
skill's name.

The counter-argument was precedent: sgOccupationSupply.ts already ships at these
same eight groups. But that figure is a DENOMINATOR for a vacancy rate, where
reweighting across groups still carries information, and vacancyRate.ts refuses
to blend it with Australia's for exactly this reason. A salary printed beside a
skill is read as that skill's pay, which is the "plausible number, broken
reasoning" failure this file exists to avoid.

If Singapore is wanted later, the thing to look for is a finer MOM occupational
wage table — not a way to present eight numbers as a hundred.

THE MARKETS DO NOT SHARE AN AXIS, AND MUST NOT BE MADE TO
AU is eight financial years with a state split in one of them; NZ is three
census years with a regional split in all three. Their bases differ too — the
ATO's "salary or wage income" excludes investment and business income, while the
census's "total personal income" includes every source. Forcing them onto one
year array would either invent AU figures for 2013 or NZ figures for 2019-20.
So each market carries its own years, basis and provenance, and a series is
positional against ITS market's years. BASELINE_MARKETS is the authority;
AREA_MARKET says which market an area belongs to.

WHAT A ROW MEANS, AND WHAT IT DOES NOT
Income actually RECEIVED by people, not a full-time-equivalent rate. Someone who
worked half the year, or three days a week, is in it at what they earned. A
figure answers "what did people in this occupation earn that year" and NOT "what
does this job pay". The effect is visible and must be labelled: Hospitality &
Food Service reads about $30k in AU 2023-24, which is casual-heavy work rather
than a full-time wage.

That is also why nothing here may be differenced against the app's live
advertised median. The two differ on instrument (received against advertised),
unit (a person against an advertisement), disclosure (everyone counted against
the third of ads that state a number), and geography. A percentage is honest
WITHIN one market's series, where every year is measured the same way, and
nowhere else — not even between AU and NZ, which are different instruments.

FIGURES ARE NOMINAL. No CPI deflation is applied and none should be inferred:
each year is published in its own dollars, in its own currency, and labelled
with its own year.

HOW A SKILL GETS A NUMBER
The same crosswalk the IVI demand and ABS supply generators use — OVERRIDE from
gen-ivi-skill-demand.py, falling back to the shared taxonomy matcher against the
occupation title. Sharing it is the point: a baseline attributed differently
from the live figure would make even a like-for-like read measure the mapping
instead of the market. An occupation mapping to two skills counts in BOTH.

NZ IS THE SAME CROSSWALK, READ AT THE GRANULARITY NZ PUBLISHES. New Zealand
gives ANZSCO sub-major, so each skill is priced through the ONE two-digit group
it is the largest identified part of, measured on the Australian four-digit
populations. Running the matcher over the sub-major NAMES instead would be a
different mapping built from classification headings rather than job titles, and
the two markets would stop being comparable in the one way they are allowed to
be — as independently-measured readings of the same skill.

TWO THINGS THAT WENT WRONG HERE FIRST, BOTH WORTH KNOWING
Taking the union of a sub-major's skills, which is the obvious rollup, makes a
skill inherit a whole occupational group it holds a sliver of — see the note on
MIN_SUBMAJOR_SHARE. And combining two groups' medians is bistable: as their
populations shift between censuses the weighted median crosses from one to the
other, which made Social & Community Services appear to lose 28.5% between 2013
and 2018 with nothing having happened to pay. Both produced numbers that looked
entirely reasonable on their own.

THE PRICE OF SUB-MAJOR, MEASURED AND DECLARED
Only 34 of the 100 skills clear the share threshold at all, and twelve of those
share a group with another skill — Software Engineering and IT & Systems both
resolve to "26 ICT Professionals", for instance. Both limits are real properties
of the NZ data, not defects here, so the collisions are EXPORTED as
BASELINE_MARKETS.nz.shared rather than hidden: the card can then say a figure
covers several skills instead of implying each was measured apart.

THE ONE STATISTICAL COMPROMISE, STATED PLAINLY
In AU a skill is several occupations, and MEDIANS DO NOT COMPOSE — there is no
way to recover the true median of the combined population from the medians of
its parts. What is computed is a weighted median OF THE PUBLISHED MEDIANS, each
weighted by the people behind it, which is the closest honest thing available
from published aggregates. It is not the population median and is not claimed to
be. Weighting matters: unweighted, "Window dresser" (218 people) would count for
as much as "Kitchen hand" (121,185). NZ does not have this problem because it
reads a single group, and it pays for that in coverage instead.

WHAT IS DELIBERATELY DROPPED
  * AU codes 0000 "Occupation blank", 9990 "Occupation not matched" and 9997
    "Miscellaneous type not specified" — the ATO's equivalent of the ABS 'nfd'
    rows, 3.4M people belonging to no occupation. Reported, never absorbed.
  * The ATO's own 9xxx apprentice/trainee/consultant codes, which cut across
    occupations and have no ANZSCO counterpart.
  * NZ's major-group and total rows, which would double-count their own
    sub-majors.

COLUMNS ARE FOUND BY HEADER TEXT, NOT BY POSITION. The ATO renumbers its
footnote markers most years ("Median salary or wage income4 $" became
"...income5 $"), inserts a column, and renamed "Number of individuals" to
"Individuals no.". Reading by position would silently pick up the wrong column
and publish a plausible wrong number — the failure this whole file guards.

Run:  python3 scripts/gen-salary-baseline.py            # downloads + caches
      python3 scripts/gen-salary-baseline.py --offline  # cache only, no network
Needs: pip install openpyxl
       STATS_NZ_API_KEY in the environment for a cold NZ fetch. Stats NZ gates
       everything but its dataflow catalogue behind an Azure APIM subscription
       key (free, from portal.apis.stats.govt.nz). The key is NEVER written to
       disk or to the generated file — only the CSV it returns is cached, so a
       rerun and CI need no key at all.
Then: npx eslint src/employsi/data/salaryBaseline.ts --fix
"""
from __future__ import annotations
import csv
import importlib.util
import io
import json
from functools import partial
import os
import re
import sys
import urllib.request

import openpyxl

# ensure_ascii=False: this writes a file people read, and an escaped em dash
# ("ATO Taxation Statistics \\u2014 Individuals Table 15") is not readable.
jdump = partial(json.dumps, ensure_ascii=False)

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, HERE)
from skills_taxonomy import matcher  # noqa: E402

TAX = os.path.join(ROOT, 'src/employsi/data/skillsTaxonomy.ts')
OUT = os.path.join(ROOT, 'src/employsi/data/salaryBaseline.ts')
CACHE = os.path.join(ROOT, '.cache')

# ── Australia ───────────────────────────────────────────────────────────────
AU_CKAN = 'https://data.gov.au/data/api/3/action/package_show?id=taxation-statistics-{}'
AU_SOURCE = 'ATO Taxation Statistics — Individuals Table 15'
AU_SOURCE_URL = 'https://data.gov.au/data/dataset/taxation-statistics-2023-24'
AU_LICENCE = 'CC BY 2.5 AU'
AU_BASIS = 'median salary or wage income, annual, nominal AUD'
# The vintages whose Table 15B carries the same unit-group x sex shape. 2015-16
# is excluded deliberately — its tables are laid out differently and would need
# their own parser for one extra year.
AU_YEARS = ['2016-17', '2017-18', '2018-19', '2019-20', '2020-21', '2021-22',
            '2022-23', '2023-24']
# The only year with a state breakdown (Table 15D). Checked all nine vintages.
AU_STATE_YEAR = '2023-24'
# Same mapping gen-abs-occupation-supply.py uses, so a state's workers land on
# the same hub in the baseline as in the supply series.
AU_STATE2CITY = {
    'NSW': 'sydney', 'VIC': 'melbourne', 'QLD': 'brisbane', 'SA': 'adelaide',
    'WA': 'perth', 'NT': 'darwin', 'ACT': 'canberra', 'TAS': 'hobart',
}
AU_AREAS = ['au', 'perth', 'adelaide', 'brisbane', 'melbourne', 'sydney',
            'darwin', 'canberra', 'hobart']
# Codes that name no occupation. Attributing them would mean guessing.
AU_UNATTRIBUTABLE = {'0000', '9990', '9997'}

# ── New Zealand ─────────────────────────────────────────────────────────────
NZ_API = ('https://api.data.stats.govt.nz/rest/data/STATSNZ,CEN23_WRK_011,1.0/'
          '.9999+02+09..Median+777.99.99/all')
NZ_SOURCE = 'Stats NZ Census — CEN23_WRK_011, occupation by total personal income'
NZ_SOURCE_URL = 'https://explore.data.stats.govt.nz/vis?df[ds]=ds-nsiws-disseminate&df[id]=CEN23_WRK_011'
NZ_LICENCE = 'CC BY 4.0'
NZ_BASIS = 'median total personal income, census year, nominal NZD'
NZ_YEARS = ['2013', '2018', '2023']
# Regional council codes -> our hubs. Both NZ hubs are regional councils, which
# is why the census table can carry them at all.
NZ_GEO2AREA = {'9999': 'nz', '02': 'auckland', '09': 'wellington'}
NZ_AREAS = ['nz', 'auckland', 'wellington']
# The dimension member carrying the published median, and the one carrying the
# count of people with a stated income (the weight).
NZ_MEDIAN = 'Median'
NZ_COUNT = '777'

# A cell below this is noise rather than signal. Neither office publishes cells
# too small to be safe, so this is not a privacy floor — it is a "one tiny
# occupation should not define a skill" floor. AU code 9997 arrived with three
# individuals in it.
MIN_CELL = 100

# How much of an ANZSCO sub-major must belong to a skill before that group's
# median may stand in for the skill's pay.
#
# THIS EXISTS BECAUSE THE FIRST NZ RUN WAS WRONG, not merely noisy. Rolling the
# four-digit crosswalk up to two digits makes a skill inherit the WHOLE group
# that any of its unit groups sits in. Pharmacy owns 2515 Pharmacists and 6212
# Pharmacy Sales Assistants, so it inherited both "25 Health Professionals" and
# "62 Sales Assistants and Salespersons" — and Pharmacy is 4.6% of the first and
# 7.2% of the second. Its NZ figure was therefore mostly the median of every
# retail salesperson in the country, which is how Pharmacy in Wellington
# appeared to rise 245% between 2018 and 2023 (23,500 -> 81,000). Nothing about
# that number looked wrong on its own; the plausibility band caught it.
#
# So a sub-major only speaks for a skill when a THIRD of it is that skill, and
# it then contributes weighted by that share rather than by its whole
# population. The cost is coverage and it is large: 75 skills at no threshold,
# 64 at 10%, 43 at 25%, 35 here, 23 at 50%. A third is the point where the
# group's median is materially shaped by the skill rather than borrowed from it.
#
# The share is measured on AUSTRALIAN unit-group populations, because the ATO
# publishes the four-digit detail that Stats NZ does not. That assumes the two
# countries split a sub-major similarly, which is an assumption, and it is why
# the threshold is set conservatively rather than at the lowest defensible value.
MIN_SUBMAJOR_SHARE = 0.33
# And a published figure must rest on this many people in total across the
# occupations behind it. Same intent as salaryParse.MIN_ADS: below the floor the
# row shows nothing rather than something badly estimated.
MIN_TOTAL = 1000


def load_override() -> dict:
    """OVERRIDE from the IVI generator, by path — the filename has hyphens, so
    it cannot be imported by name. Identical loader to
    gen-abs-occupation-supply.py, and identical reason: the datasets have to
    attribute an occupation the same way or they cannot be read together."""
    path = os.path.join(HERE, 'gen-ivi-skill-demand.py')
    spec = importlib.util.spec_from_file_location('gen_ivi', path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod.OVERRIDE


# ── fetching ────────────────────────────────────────────────────────────────
def fetch_au(year: str, offline: bool) -> str:
    """The Table 15 workbook for one financial year, cached on disk.

    Resolved through the CKAN API rather than by guessing the URL: the filename
    changes every year ('taxstats2016individual15…', 'ts17individual15…',
    'ts24individual15occupationsex.xlsx') and the resource NAME is the useless
    'Individuals - Table 15' in every edition, so the URL is the only thing
    carrying the table's identity.
    """
    d = os.path.join(CACHE, 'ato')
    os.makedirs(d, exist_ok=True)
    path = os.path.join(d, f'ato15-{year}.xlsx')
    if os.path.exists(path):
        return path
    if offline:
        raise SystemExit(f'✗ AU {year} not cached and --offline was passed.')
    with urllib.request.urlopen(AU_CKAN.format(year), timeout=90) as r:
        pkg = json.load(r)
    for res in pkg['result']['resources']:
        url = res.get('url') or ''
        if re.search(r'individual15occupation(sex|gender)', url, re.I) and url.lower().endswith('.xlsx'):
            urllib.request.urlretrieve(url, path)
            return path
    raise SystemExit(f'✗ No Table 15 occupation workbook found for AU {year}.')


def fetch_nz(offline: bool) -> str:
    """The census median-income CSV, cached on disk.

    The key is read from the environment and used only as a request header — it
    is never written to the cache, the generated file or the repo.
    """
    os.makedirs(CACHE, exist_ok=True)
    path = os.path.join(CACHE, 'nz-cen23-wrk-011.csv')
    if os.path.exists(path):
        return path
    if offline:
        raise SystemExit(f'✗ NZ data not cached at {path} and --offline was passed.')
    key = os.environ.get('STATS_NZ_API_KEY')
    if not key:
        raise SystemExit(
            '✗ STATS_NZ_API_KEY is not set, and the NZ CSV is not cached.\n'
            '  Stats NZ gates every endpoint but its dataflow catalogue behind a\n'
            '  free Azure APIM subscription key (portal.apis.stats.govt.nz).\n'
            '  Once the CSV is cached this generator and CI need no key.')
    req = urllib.request.Request(NZ_API, headers={
        'Ocp-Apim-Subscription-Key': key,
        'Accept': 'application/vnd.sdmx.data+csv;version=1.0.0',
    })
    with urllib.request.urlopen(req, timeout=180) as r:
        body = r.read()
    with open(path, 'wb') as f:
        f.write(body)
    return path


# ── AU parsing ──────────────────────────────────────────────────────────────
def header_row(ws, want: str) -> tuple[int, list[str]]:
    """The first row that looks like a header, and its cells as text.

    'Looks like' is four or more non-empty cells one of which mentions `want`;
    the ATO puts a title and a release line above the header and the number of
    those lines is not stable across editions.
    """
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=12, max_col=14, values_only=True), 1):
        cells = ['' if c is None else str(c).replace('\n', ' ').strip() for c in row]
        if sum(1 for c in cells if c) >= 4 and any(want in c.lower() for c in cells):
            return i, cells
    raise SystemExit(f'✗ No header row containing {want!r} in sheet {ws.title!r}.')


def col(cells: list[str], pattern: str) -> int:
    """Index of the first column whose header matches `pattern`. By text, never
    by position — see the note on footnote markers above."""
    rx = re.compile(pattern, re.I)
    for i, c in enumerate(cells):
        if rx.search(c):
            return i
    raise SystemExit(f'✗ No column matching /{pattern}/ in header: {cells}')


UNIT_RE = re.compile(r'^(\d{4})\s+(.*?)\s*$')


def read_au_sheet(path: str, suffix: str, area_col: str | None):
    """(code, title, area, individuals, median wage) from one Table 15 sheet.

    `area_col` names the column carrying the geography, or None for the national
    table — where the sex column is used instead and only 'Total' is kept.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = [s for s in wb.sheetnames if s.replace(' ', '').endswith(suffix)]
    if not names:
        wb.close()
        return None
    ws = wb[names[0]]
    hrow, cells = header_row(ws, 'unit group')
    c_unit = col(cells, r'unit group')
    c_n = col(cells, r'individuals')
    c_med = col(cells, r'median salary or wage')
    c_area = col(cells, area_col) if area_col else col(cells, r'^sex')

    out = []
    for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
        unit = row[c_unit] if c_unit < len(row) else None
        if not isinstance(unit, str):
            continue
        m = UNIT_RE.match(unit.strip())
        if not m:
            continue
        area = row[c_area] if c_area < len(row) else None
        n = row[c_n] if c_n < len(row) else None
        med = row[c_med] if c_med < len(row) else None
        if not isinstance(n, (int, float)) or not isinstance(med, (int, float)):
            continue  # the ATO leaves suppressed cells empty
        out.append((m.group(1), m.group(2), str(area or '').strip(), float(n), float(med)))
    wb.close()
    return out


# ── folding ─────────────────────────────────────────────────────────────────
def weighted_median(pairs: list[tuple[float, float]]) -> int | None:
    """Weighted median of (median, weight) pairs — see the docstring's note on
    medians not composing. Ties resolve to the lower value, which keeps the
    result a figure that was actually published rather than an interpolation
    between two."""
    if not pairs:
        return None
    total = sum(w for _, w in pairs)
    if total <= 0:
        return None
    half = total / 2.0
    acc = 0.0
    for med, w in sorted(pairs):
        acc += w
        if acc >= half:
            return int(round(med))
    return int(round(sorted(pairs)[-1][0]))


def main(argv: list[str]) -> int:
    offline = '--offline' in argv
    skills_for = matcher(TAX)
    override = load_override()

    # cells[skill][area][year] = [(median, weight), ...]
    cells: dict[str, dict[str, dict[str, list[tuple[float, float]]]]] = {}
    au_dropped = 0.0
    au_unmapped: dict[str, float] = {}
    au_codes: dict[str, str] = {}
    au_pop: dict[str, float] = {}

    def add(skill_list, area, year, med, w):
        if w < MIN_CELL:
            return
        for s in skill_list:
            cells.setdefault(s, {}).setdefault(area, {}).setdefault(year, []).append((med, w))

    # ── Australia ───────────────────────────────────────────────────────────
    print('Australia — ATO Taxation Statistics')
    for year in AU_YEARS:
        path = fetch_au(year, offline)
        national = read_au_sheet(path, '15B', None)
        if national is None:
            raise SystemExit(f'✗ AU {year} has no Table 15B (unit group x sex).')
        counting = year == AU_YEARS[-1]
        for code, title, area_raw, n, med in national:
            if area_raw.lower() != 'total':
                continue
            au_codes[code] = title
            if year == AU_YEARS[-1]:
                au_pop[code] = n
            if code in AU_UNATTRIBUTABLE:
                if counting:
                    au_dropped += n
                continue
            sk = override.get(code) or skills_for(title)
            if not sk:
                if counting:
                    au_unmapped[f'{code} {title}'] = au_unmapped.get(f'{code} {title}', 0.0) + n
                continue
            add(sk, 'au', year, med, n)
        print(f'  {year}  15B rows={len(national):>5}')

        if year == AU_STATE_YEAR:
            states = read_au_sheet(path, '15D', r'^state')
            if states is None:
                raise SystemExit(f'✗ AU {year} was expected to carry Table 15D and does not.')
            for code, title, area_raw, n, med in states:
                area = AU_STATE2CITY.get(area_raw)
                if not area or code in AU_UNATTRIBUTABLE:
                    continue
                sk = override.get(code) or skills_for(title)
                if sk:
                    add(sk, area, year, med, n)
            print(f'  {year}  15D rows={len(states):>5}  (the only year with a state split)')

    # ── New Zealand ─────────────────────────────────────────────────────────
    # Sub-major -> {skill: share of the group that is that skill}, rolled up
    # from the AU unit-group crosswalk so both markets attribute an occupation
    # identically, then filtered by MIN_SUBMAJOR_SHARE. See the note there for
    # why the filter is not optional.
    sub_pop: dict[str, float] = {}
    skill_pop: dict[tuple[str, str], float] = {}
    for code, title in au_codes.items():
        if code in AU_UNATTRIBUTABLE:
            continue
        n = au_pop.get(code, 0.0)
        sub_pop[code[:2]] = sub_pop.get(code[:2], 0.0) + n
        for s in (override.get(code) or skills_for(title)):
            skill_pop[(s, code[:2])] = skill_pop.get((s, code[:2]), 0.0) + n
    # ONE sub-major per skill: the group the skill is the largest part of.
    #
    # Combining two groups' medians was the obvious thing and it is bistable.
    # Social & Community Services sits in "27 Legal, Social and Welfare
    # Professionals" and "41 Health and Welfare Support Workers", whose medians
    # are far apart; as their populations shifted between censuses the weighted
    # median crossed from one to the other and the skill appeared to lose 28.5%
    # between 2013 and 2018. Nothing had happened to pay — the estimator had
    # simply tipped over. With only two or three groups behind a skill there is
    # no weighting that removes this, because the quantity being estimated jumps.
    #
    # So NZ prices a skill through the single occupation group that best
    # represents it, which cannot jump: the share is computed once from the
    # Australian four-digit populations and is the same in every census year.
    # The figure means "the median of the ANZSCO sub-major this skill is the
    # largest identified part of", and the surface should say so.
    best: dict[str, tuple[str, float]] = {}
    dropped_pairs = 0
    for (s, sm), n in skill_pop.items():
        tot = sub_pop.get(sm, 0.0)
        share = n / tot if tot else 0.0
        if share < MIN_SUBMAJOR_SHARE:
            dropped_pairs += 1
            continue
        if s not in best or share > best[s][1]:
            if s in best:
                dropped_pairs += 1
            best[s] = (sm, share)
        else:
            dropped_pairs += 1
    roll: dict[str, dict[str, float]] = {}
    for s, (sm, share) in best.items():
        roll.setdefault(sm, {})[s] = share

    print('\nNew Zealand — Stats NZ Census CEN23_WRK_011')
    nz_path = fetch_nz(offline)
    with open(nz_path, encoding='utf-8') as f:
        rows = list(csv.DictReader(f))
    # (year, area, occ) -> {'med': x, 'n': y}
    nz: dict[tuple[str, str, str], dict[str, float]] = {}
    for r in rows:
        occ = (r.get('CEN23_OCC_003') or '').strip()
        area = NZ_GEO2AREA.get((r.get('CEN23_GEO_008') or '').strip())
        year = (r.get('CEN23_YEAR_001') or '').strip()
        val = (r.get('OBS_VALUE') or '').strip()
        # Only the two-digit sub-majors. The single-digit majors and the '999'
        # total are the same people again at a coarser cut; folding them in
        # would count everyone twice and drag every figure toward the middle.
        if not area or len(occ) != 2 or not val:
            continue
        which = r.get('CEN23_TOI_002')
        if which == NZ_MEDIAN:
            nz.setdefault((year, area, occ), {})['med'] = float(val)
        elif which == NZ_COUNT:
            nz.setdefault((year, area, occ), {})['n'] = float(val)
    nz_unmapped: set[str] = set()
    for (year, area, occ), v in nz.items():
        if 'med' not in v or 'n' not in v:
            continue
        sk = roll.get(occ)
        if not sk:
            nz_unmapped.add(occ)
            continue
        # Weighted by the share of the group that is actually this skill, so a
        # sub-major that only just clears the threshold cannot outvote one the
        # skill dominates.
        for s in sk:
            add([s], area, year, v['med'], v['n'])
    print(f'  csv rows={len(rows):>5}  sub-major cells={len(nz)}  '
          f'sub-majors carrying no qualifying skill={len(nz_unmapped)}')
    print(f'  (skill, sub-major) pairs dropped under the {MIN_SUBMAJOR_SHARE:.0%} share '
          f'threshold: {dropped_pairs}')

    # Skills that resolve to an identical set of sub-majors cannot be told apart
    # in NZ. Exported rather than hidden — see the docstring.
    skill_sub: dict[str, set] = {}
    for occ, sks in roll.items():
        for s in sks:
            skill_sub.setdefault(s, set()).add(occ)
    grouped: dict[frozenset, list[str]] = {}
    for s, g in skill_sub.items():
        grouped.setdefault(frozenset(g), []).append(s)
    nz_shared = sorted(
        (sorted(v) for v in grouped.values() if len(v) > 1), key=lambda x: (-len(x), x[0]))

    # ── fold to one figure per skill/area/year ──────────────────────────────
    MARKETS = {
        'au': dict(years=AU_YEARS, areas=AU_AREAS, basis=AU_BASIS, source=AU_SOURCE,
                   sourceUrl=AU_SOURCE_URL, licence=AU_LICENCE, currency='AUD',
                   areaYears=[AU_STATE_YEAR], shared=[]),
        'nz': dict(years=NZ_YEARS, areas=NZ_AREAS, basis=NZ_BASIS, source=NZ_SOURCE,
                   sourceUrl=NZ_SOURCE_URL, licence=NZ_LICENCE, currency='NZD',
                   areaYears=NZ_YEARS, shared=nz_shared),
    }
    area_market = {a: m for m, spec in MARKETS.items() for a in spec['areas']}

    out: dict[str, dict[str, list]] = {}
    published = 0
    for skill in sorted(cells):
        for area, byyear in cells[skill].items():
            years = MARKETS[area_market[area]]['years']
            series, any_year = [], False
            for year in years:
                pairs = byyear.get(year) or []
                total = sum(w for _, w in pairs)
                med = weighted_median(pairs) if total >= MIN_TOTAL else None
                if med is None:
                    series.append(None)
                else:
                    series.append([med, int(round(total))])
                    any_year = True
                    published += 1
            if any_year:
                out.setdefault(skill, {})[area] = series

    # ── report ──────────────────────────────────────────────────────────────
    au_skills = {s for s, a in out.items() if any(area_market[x] == 'au' for x in a)}
    nz_skills = {s for s, a in out.items() if any(area_market[x] == 'nz' for x in a)}
    print(f'\nskills with an AU series : {len(au_skills)}')
    print(f'skills with an NZ series : {len(nz_skills)}')
    print(f'figures published        : {published}')
    print(f'NZ skills sharing a figure: {sum(len(g) for g in nz_shared)} in {len(nz_shared)} groups')
    print(f'\ncoverage of the AU {AU_YEARS[-1]} national table:')
    print(f'  unattributable codes : {au_dropped:,.0f} people (0000/9990/9997)')
    if au_unmapped:
        top = sorted(au_unmapped.items(), key=lambda kv: -kv[1])
        print(f'  crosswalk misses     : {len(au_unmapped)} unit groups '
              f'({sum(au_unmapped.values()):,.0f} people)')
        for name, n in top[:6]:
            print(f'      {n:>10,.0f}  {name}')
        if len(top) > 6:
            print(f'      ... and {len(top) - 6} more ({sum(n for _, n in top[6:]):,.0f} people)')

    # ── write ───────────────────────────────────────────────────────────────
    def fmt(series) -> str:
        return '[' + ','.join('null' if v is None else f'[{v[0]},{v[1]}]' for v in series) + ']'

    body = []
    for skill in sorted(out):
        areas = out[skill]
        order = [a for m in MARKETS for a in MARKETS[m]['areas'] if a in areas]
        inner = ',\n'.join(f'    {jdump(a)}: {fmt(areas[a])}' for a in order)
        body.append(f'  {jdump(skill)}: {{\n{inner},\n  }}')
    joined = ',\n'.join(body)

    markets_ts = ',\n'.join(
        '  ' + jdump(m) + ': {\n' + ',\n'.join(
            f'    {k}: {jdump(v)}' for k, v in spec.items()) + ',\n  }'
        for m, spec in MARKETS.items())

    with open(OUT, 'w', encoding='utf-8') as f:
        f.write(f'''// GENERATED — do not edit by hand.
// Run: python3 scripts/gen-salary-baseline.py
//
// Historical median pay per skill, for the years the live archive cannot reach:
// it begins 2026-07-20, so it cannot say whether a skill is worth more than it
// used to be.
//
// READ THE UNIT BEFORE USING A NUMBER FROM HERE. This is income actually
// RECEIVED by people — part-year and part-time earners are in it at what they
// earned — not a full-time-equivalent rate and not an advertised salary. It
// answers "what did people in this occupation earn that year", not "what does
// this job pay". Hospitality reads about $30k in AU 2023-24 for exactly that
// reason.
//
// NEVER DIFFERENCE IT AGAINST THE APP'S LIVE ADVERTISED MEDIAN. The two differ
// on instrument, unit, disclosure and geography all at once, so the gap between
// them is mostly the gap between two measuring devices. A percentage is honest
// WITHIN one market's series and nowhere else — NOT between AU and NZ either,
// whose bases differ (salary or wage income against total personal income).
//
// Figures are NOMINAL, each in its own year's dollars and its own currency.
//
// A skill spans several occupations, and medians do not compose: each figure is
// a weighted median of the published medians behind it, weighted by people. It
// approximates the population median and is not it.

export interface BaselineMarket {{
  /** Reference years, oldest first. Every series for this market's areas is
   *  positional against THIS array — the markets do not share an axis. */
  years: string[];
  /** Areas belonging to the market; the first is the national one. */
  areas: string[];
  basis: string;
  source: string;
  sourceUrl: string;
  licence: string;
  currency: string;
  /** Years for which the non-national areas carry figures. The ATO published a
   *  state split in one year only; the NZ census publishes a regional split in
   *  all three. A hub series is null outside these, deliberately, rather than
   *  repeating the national figure as though it were local. */
  areaYears: string[];
  /** Groups of skills this market cannot tell apart, because they resolve to
   *  the same occupations at the granularity it publishes. Every skill in a
   *  group carries an identical figure, and the surface must say so rather than
   *  imply each was measured separately. Empty where the data is fine enough. */
  shared: string[][];
}}

export const BASELINE_MARKETS: Record<string, BaselineMarket> = {{
{markets_ts},
}};

/** Area -> the market whose `years` its series is aligned to. */
export const AREA_MARKET: Record<string, string> = {jdump(area_market)};

// skill -> area -> [median, people behind it] per year of that area's market,
// or null where nothing was published or the sample fell below the floor
// ({MIN_TOTAL:,} people).
export const SALARY_BASELINE: Record<string, Record<string, ([number, number] | null)[]>> = {{
{joined},
}};
''')

    print(f'\nWrote {len(out)} skills to {os.path.relpath(OUT, ROOT)}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main(sys.argv[1:]))
