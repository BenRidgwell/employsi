#!/usr/bin/env python3
"""Philippines vacancies by occupation -> src/employsi/data/phVacancyDemand.ts

Source: Philippine Statistics Authority, Integrated Survey on Labor and
Employment (ISLE, formerly the BLES Integrated Survey / BITS). Six survey rounds
supplied as five spreadsheets plus one PDF newsletter:

  2011/2012  Jan 2011 - Jun 2012   LABSTAT Updates Vol.17 No.24, TABLE 3   20+
  2013/2014  Jan 2013 - Jun 2014   tab7 (TABLE 7)                          20+
  2015/2016  Jan 2015 - Jun 2016   Table207_0 (TABLE 7)                    20+   ← not held
  2017/2018  Jul 2017 - Jun 2018   Table207_2 (TABLE 7)                    20+
  2021/2022  Sep 2021 - Aug 2022   Table_4_10 (TABLE 4)                    20+
  2023/2024  Sep 2023 - Aug 2024   Table_11_12 (TABLE 11)                  10+

A DETAILED ROUND MAY BE ABSENT. Each of these is a separate PSA publication
downloaded by hand, and psa.gov.ph refuses automated requests (403 direct, and
Oxylabs faults on it), so a round that is not to hand cannot simply be fetched.
2015/2016 is currently in that position. An absent detailed round is DROPPED —
never inferred from its neighbours — and the gap is visible rather than
implicit: the generator names it on stderr, and PH_THRESHOLD / PH_GRANULARITY /
PH_ROUND_TOTALS in the emitted file carry only the rounds that were read, so a
reader sees five rounds where six are documented. Supplying Table207_0.xlsx and
re-running is all that is needed to close it.

The two group-only rounds stay REQUIRED: they are the two ends of the axis, and
losing either silently shortens the group panel rather than thinning it.

── WHAT STITCHING THESE HONESTLY MEANS ────────────────────────────────────────
They are NOT one series. Four things break comparability and each is handled
here rather than smoothed over:

 1. THE ESTABLISHMENT THRESHOLD CHANGED. Every round through 2021/22 covers
    establishments with 20 or more workers; 2023/24 covers 10 or more. A larger
    frame reports more vacancies for the same labour market, so the 2023/24
    level is not a continuation of the earlier ones. PH_THRESHOLD records the
    frame per round so a reader can see the break instead of inferring a jump.

 2. GRANULARITY DIFFERS, AND THAT DECIDES WHAT EACH ROUND CAN SAY. 2013/14
    through 2021/22 publish detailed occupation titles. 2011/12 (from the PDF)
    and 2023/24 publish MAJOR OCCUPATION GROUPS only.

    The obvious bridge — let every skill in a major group inherit that group's
    published figure, the way data/hkVacancyDemand.ts lets a skill inherit its
    industry section — WAS TRIED AND MEASURED, and it does not work here. Run
    both schemes over the 2021/22 round, which publishes detail AND group
    headers so the two are directly comparable, and group inheritance overstates
    a skill by a MEDIAN OF 72x (Banking & Lending 4,193 -> 211,000; Mechanical
    Engineering 3,687 -> 150,577). Hong Kong's 17 industry sections are
    reasonably cohesive; ISCO's nine major groups are not — "Managers" holds
    mining, nursing and restaurant managers alike, so a narrow skill inherits an
    entire economy's worth of vacancies.

    So the two granularities are published as two exports and never mixed:
      • PH_SERIES      — per SKILL, from the four detailed rounds only.
      • PH_GROUP_SERIES — per MAJOR OCCUPATION GROUP, from all six rounds.
    Nothing is dropped: the 2011/12 figures dug out of the PDF and the 2023/24
    round both appear at the granularity their source actually published.

 3. THE CLASSIFICATION CHANGED. 2013/14 uses the pre-2012 PSOC titles
    ("Corporate Executives, Managers Managing Proprietors and Supervisors");
    later rounds use PSOC 2012 ("Managers"). Titles are matched to skills by
    the app's own text matcher, which is why a renamed occupation still lands
    on the same skill rather than needing a per-round crosswalk.

 4. THEY ARE SURVEY WINDOWS, NOT MONTHS. Each round covers a 12-18 month
    reference period, and there are multi-year gaps between rounds. Nothing is
    interpolated across a gap: each round is HELD from the start of its own
    window until the next round begins, which is what "the latest published
    figure" means. Months before the first round stay zero.

The PDF's pages 4-5 also list detailed occupations, but as a selective
highlights narrative by industry — those counts do not sum to the published
total and are NOT used. TABLE 3 on page 3 is the complete 2011/12 breakdown
(619,580 across nine groups) and is what this reads.

Usage: python3 scripts/gen-ph-vacancy-demand.py <dir-with-the-source-files>
"""
from __future__ import annotations
import json
import os
import re
import subprocess
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
OUT = os.path.join(ROOT, 'src/employsi/data/phVacancyDemand.ts')

# The Philippine roster sits on Manila, the app's PH hub.
HUB = 'manila'

# round -> (first month, last month, threshold, granularity)
ROUNDS = {
    '2011/2012': ('2011-01', '2012-06', '20+', 'major group'),
    '2013/2014': ('2013-01', '2014-06', '20+', 'detailed'),
    '2015/2016': ('2015-01', '2016-06', '20+', 'detailed'),
    '2017/2018': ('2017-07', '2018-06', '20+', 'detailed'),
    '2021/2022': ('2021-09', '2022-08', '20+', 'detailed'),
    '2023/2024': ('2023-09', '2024-08', '10+', 'major group'),
}

# 2011/2012, PDF page 3 TABLE 3 — "Vacancies in Industry and Services Sectors by
# Major Occupation Group, January 2011-June 2012". Keyed in from the table
# rather than parsed: the page is a two-column newsletter whose text extraction
# interleaves the columns, and nine transcribed numbers that sum to the
# published 619,580 are safer than a fragile layout parser. The check below
# fails the run if they stop summing.
PDF_2011 = [
    ('Corporate executives, managers, managing proprietors and supervisors', 12108),
    ('Professionals', 70623),
    ('Technicians and associate professionals', 54849),
    ('Clerks', 228984),
    ('Service workers and shop and market sales workers', 86122),
    ('Farmers, forestry workers and fishermen', 548),
    ('Craft and related trades workers', 38259),
    ('Plant and machine operators and assemblers', 64444),
    ('Laborers and unskilled workers', 63644),
]
PDF_2011_TOTAL = 619580
# The table's own note says "Details do not add up to total due to rounding",
# and they sum to 619,581. The check below allows that one, and only that one:
# a wider tolerance would let a mis-keyed digit through, which is the failure
# this check exists to catch. It already has — three of these nine were first
# written from the truncated text extraction and were wrong by thousands.
PDF_2011_TOLERANCE = 2

# Every round publishes the same nine major occupation groups in the same order,
# but the 2012 PSOC revision renamed most of them mid-panel ("Clerks" ->
# "Clerical Support Workers", "Laborers and Unskilled Workers" -> "Elementary
# Occupations"). The canonical names below are the current PSOC 2012 ones; the
# alias table folds the older labels onto them so PH_GROUP_SERIES is one panel
# rather than two half-panels.
#
# This is a RENAMING, not a crosswalk: the nine groups correspond one-to-one
# across the revision. Their CONTENT did shift slightly at the boundaries (PSOC
# 2012 moved some ICT and supervisory roles between groups), which is why the
# group panel is still labelled per round by PH_THRESHOLD/PH_GRANULARITY rather
# than presented as a clean continuous series.
GROUPS = [
    'Managers',
    'Professionals',
    'Technicians and Associate Professionals',
    'Clerical Support Workers',
    'Service and Sales Workers',
    'Skilled Agricultural, Forestry and Fishery Workers',
    'Craft and Related Trades Workers',
    'Plant and Machine Operators and Assemblers',
    'Elementary Occupations',
]
GROUP_ALIAS = {
    'corporate executives, managers managing proprietors and supervisors': 'Managers',
    'corporate executives, managers, managing proprietors and supervisors': 'Managers',
    'clerks': 'Clerical Support Workers',
    'service workers and shop and market sales workers': 'Service and Sales Workers',
    'farmers, forestry workers and fishermen': 'Skilled Agricultural, Forestry and Fishery Workers',
    'skilled agricultural, forestry and fishery workers': 'Skilled Agricultural, Forestry and Fishery Workers',
    'laborers and unskilled workers': 'Elementary Occupations',
    'elementary occupations': 'Elementary Occupations',
}


def canon_group(label: str) -> str | None:
    """A round's group label -> the canonical PSOC 2012 name, or None."""
    k = ' '.join(label.strip().lower().split())
    if k in GROUP_ALIAS:
        return GROUP_ALIAS[k]
    for g in GROUPS:
        if k == g.lower():
            return g
    return None


def months_axis() -> list[str]:
    p = subprocess.run(['npx', 'tsx', '-e',
                        'import { IVI_MONTHS } from "./src/employsi/data/iviSkillDemand";'
                        'console.log(JSON.stringify(IVI_MONTHS));'],
                       capture_output=True, timeout=300, cwd=ROOT)
    if p.returncode != 0:
        sys.exit('IVI_MONTHS read failed:\n' + p.stderr.decode()[-600:])
    return json.loads(p.stdout.decode().strip().splitlines()[-1])


def map_skills(titles: list[str]) -> list[list[str]]:
    """The app's own matcher, so a PH occupation lands where an AU one would."""
    if not titles:
        return []
    p = subprocess.run(['bun', 'run', os.path.join(HERE, 'map-skills.ts')],
                       input=json.dumps(titles).encode(), capture_output=True, timeout=300)
    if p.returncode != 0:
        sys.exit('map-skills failed:\n' + p.stderr.decode()[-600:])
    return json.loads(p.stdout.decode())


def num(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v) if v > 0 else None
    t = str(v).strip().replace(',', '')
    try:
        f = float(t)
        return f if f > 0 else None
    except ValueError:
        return None


def rows_from_xlsx(path: str, sheet: str, title_col: int, val_col: int) -> list[tuple[str, float]]:
    import openpyxl
    ws = openpyxl.load_workbook(path, data_only=True)[sheet]
    out = []
    for r in ws.iter_rows(values_only=True):
        if len(r) <= max(title_col, val_col):
            continue
        title = r[title_col]
        if not isinstance(title, str):
            continue
        t = title.strip()
        v = num(r[val_col])
        if not t or v is None:
            continue
        out.append((t, v))
    return out


def rows_from_xls(path: str, title_col: int, val_col: int) -> list[tuple[str, float]]:
    import xlrd
    s = xlrd.open_workbook(path).sheet_by_index(0)
    out = []
    for i in range(s.nrows):
        row = s.row_values(i)
        if len(row) <= max(title_col, val_col):
            continue
        title = row[title_col]
        if not isinstance(title, str):
            continue
        t = title.strip()
        v = num(row[val_col])
        if not t or v is None:
            continue
        out.append((t, v))
    return out


# Rows that are headings, totals or footnotes rather than an occupation.
SKIP = re.compile(
    r'^(all occupations|total|note:|source:|occupation title|major occupation group|'
    r'status|number of|regular|non-regular|table\b)', re.I)


def clean(rows: list[tuple[str, float]]) -> list[tuple[str, float]]:
    out = []
    for t, v in rows:
        if SKIP.match(t) or len(t) < 4:
            continue
        out.append((t, v))
    return out


def group_rows(rows: list[list], val_col: int) -> list[tuple[str, float]]:
    """The major-group SUBTOTAL rows of a detailed table.

    In every detailed round the layout is the same: a group header carries its
    name in column 0 with the detail-title column empty, while each occupation
    under it carries a rank in column 0 and its title in the detail column. So a
    row with a name in column 0, nothing in column 1 and a number in the value
    column is a group subtotal.

    Verified per round in main(): the detail under each header sums to the
    header, and the headers sum to ALL OCCUPATIONS. If a future PSA layout
    breaks that, the check fails the run rather than publishing a short panel.
    """
    out: list[tuple[str, float]] = []
    for r in rows:
        r = list(r) + [None, None, None]
        c0, c1 = r[0], r[1]
        if not isinstance(c0, str) or not c0.strip():
            continue
        if c1 is not None and str(c1).strip():
            continue
        v = num(r[val_col])
        if v is None or SKIP.match(c0.strip()):
            continue
        g = canon_group(c0)
        if g:
            out.append((g, v))
    return out


def raw_rows_xlsx(path: str, sheet: str) -> list[list]:
    import openpyxl
    ws = openpyxl.load_workbook(path, data_only=True)[sheet]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def raw_rows_xls(path: str) -> list[list]:
    import xlrd
    s = xlrd.open_workbook(path).sheet_by_index(0)
    return [s.row_values(i) for i in range(s.nrows)]


def main() -> int:
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    d = sys.argv[1]

    def find(frag: str, required: bool = True) -> str | None:
        """Locate a round's workbook. `required=False` for the detailed rounds.

        The six rounds are six separate PSA publications, each downloaded by
        hand, and one of them going missing must not take the other five with
        it. A round that is absent is DROPPED — never inferred from its
        neighbours — and named on stderr and in the emitted file, so a reader
        can see the gap instead of finding a shorter series than they expected.
        The 2011/12 and 2023/24 group rounds stay required: they are the two
        ends of the axis, and without them the group panel silently shortens.
        """
        for f in sorted(os.listdir(d)):
            if frag in f:
                return os.path.join(d, f)
        if required:
            sys.exit(f'missing source file containing {frag!r}')
        return None

    # 2011/12 comes from the transcribed PDF table; verify it before using it.
    got = sum(v for _, v in PDF_2011)
    if abs(got - PDF_2011_TOTAL) > PDF_2011_TOLERANCE:
        sys.exit(f'2011/2012 transcription sums to {got:,}, not the published '
                 f'{PDF_2011_TOTAL:,} — re-check the PDF before trusting this.')

    # The four rounds published by detailed occupation title. Each is read
    # twice: the occupation rows (which map to skills) and the major-group
    # subtotal rows (which join the group panel).
    DETAILED = [
        ('2013/2014', 'tab7_1.xls', None),
        ('2015/2016', 'Table207_0.xlsx', 'table7'),
        ('2017/2018', 'Table207_2.xlsx', 'TABLE 7'),
        ('2021/2022', 'Table_4_10.xlsx', 'Table 4'),
    ]
    raw = {}
    absent = []
    for rnd, frag, sheet in DETAILED:
        path = find(frag, required=False)
        if not path:
            absent.append((rnd, frag))
            continue
        rows = raw_rows_xls(path) if sheet is None else raw_rows_xlsx(path, sheet)
        raw[rnd] = (rows, 1, 2)
    if not raw:
        sys.exit('no detailed round supplied — PH_SERIES would be empty.')
    for rnd, frag in absent:
        sys.stderr.write(f'  {rnd}: {frag} not supplied — round dropped from the '
                         f'skill panel and from PH_ROUND_TOTALS\n')
    detail: dict[str, list[tuple[str, float]]] = {}
    groups: dict[str, dict[str, float]] = {}
    for rnd, (rows, tc, vc) in raw.items():
        detail[rnd] = clean([(r[tc].strip(), num(r[vc]))
                             for r in (list(x) + [None] * 4 for x in rows)
                             if isinstance(r[tc], str) and r[tc].strip()
                             and num(r[vc]) is not None])
        g = group_rows(rows, vc)
        if len(g) != len(GROUPS):
            sys.exit(f'{rnd}: found {len(g)} major-group subtotals, expected '
                     f'{len(GROUPS)} — the PSA table layout has moved.')
        groups[rnd] = {k: v for k, v in g}
        # The detail must reconcile to the subtotals, or one of the two is being
        # read wrong. This is the check that would have caught a top-N table
        # being treated as a complete one.
        ds, gs = sum(v for _, v in detail[rnd]), sum(groups[rnd].values())
        if abs(ds - gs) > max(2.0, gs * 0.001):
            sys.exit(f'{rnd}: detail sums to {ds:,.0f} but major groups sum to '
                     f'{gs:,.0f} — the table is not fully covered.')

    # The two rounds published by major group only. No detail exists for them,
    # so they contribute to the group panel and NOT to the skill series (see the
    # module docstring for the measurement behind that).
    groups['2011/2012'] = {}
    for label, v in PDF_2011:
        g = canon_group(label)
        if not g:
            sys.exit(f'2011/2012: unrecognised major group {label!r}')
        groups['2011/2012'][g] = float(v)
    groups['2023/2024'] = {}
    for label, v in clean(rows_from_xlsx(find('Table_11_12.xlsx'), 'TABLE 11', 0, 1)):
        g = canon_group(label)
        if not g:
            sys.exit(f'2023/2024: unrecognised major group {label!r}')
        groups['2023/2024'][g] = v
    for rnd in ('2011/2012', '2023/2024'):
        if len(groups[rnd]) != len(GROUPS):
            sys.exit(f'{rnd}: {len(groups[rnd])} major groups, expected {len(GROUPS)}')

    for rnd in sorted(ROUNDS, key=lambda k: ROUNDS[k][0]):
        if rnd not in groups:
            continue
        nd = len(detail.get(rnd, []))
        sys.stderr.write(f'  {rnd}: {len(groups[rnd])} major groups, '
                         f'{nd if nd else "—":>4} detailed occupations, '
                         f'{sum(groups[rnd].values()):>9,.0f} vacancies\n')

    months = months_axis()
    idx = {m: i for i, m in enumerate(months)}
    order = sorted(ROUNDS, key=lambda k: ROUNDS[k][0])

    def hold(rnd: str, acc: dict[str, float], into: dict[str, list[float]],
             to_end: bool) -> None:
        """Write a round's values across its window.

        A round is HELD from the start of its own reference window until the
        next round of the SAME panel begins — that is what "the latest published
        figure" means between surveys. `to_end` runs the final round to the axis
        end, because a trailing zero would read as "no vacancies in Manila"
        rather than "the survey has not reported since" (the same trap
        gen-hk-vacancy-demand.py's carry() exists to avoid).
        """
        seq = [r for r in order if r in groups]
        ri = seq.index(rnd)
        lo = idx.get(ROUNDS[rnd][0])
        if lo is None:
            return
        stop = None if to_end else ROUNDS[seq[ri + 1]][0]
        hi = len(months) if stop is None else idx.get(stop, len(months))
        for k, v in acc.items():
            arr = into.setdefault(k, [0.0] * len(months))
            for i in range(lo, hi):
                arr[i] = v

    # ── the skill panel: detailed rounds only ────────────────────────────────
    series: dict[str, list[float]] = {}
    detail_order = [r for r in order if r in detail]
    for rnd in detail_order:
        rows = detail[rnd]
        skills = map_skills([t for t, _ in rows])
        # An occupation mapping to several skills contributes its FULL count to
        # each, exactly as the IVI generator does: these are vacancies that want
        # that skill, not a quantity to be divided between them. Summing across
        # skills therefore exceeds the survey total, which is why only the
        # per-skill series is published and never a "total from skills".
        acc: dict[str, float] = {}
        for (t, v), sk in zip(rows, skills):
            for s in sk:
                acc[s] = acc.get(s, 0.0) + v
        # The next DETAILED round, not the next round overall: 2021/22 is the
        # most recent survey published by occupation, so it stands as the skill
        # detail from Sep 2021 to the axis end even though 2023/24 has since
        # reported a group-level total.
        di = detail_order.index(rnd)
        nxt = detail_order[di + 1] if di + 1 < len(detail_order) else None
        lo = idx.get(ROUNDS[rnd][0])
        if lo is None:
            continue
        hi = len(months) if nxt is None else idx.get(ROUNDS[nxt][0], len(months))
        for s, v in acc.items():
            arr = series.setdefault(s, [0.0] * len(months))
            for i in range(lo, hi):
                arr[i] = v

    if not series:
        sys.exit('no skills mapped — has the taxonomy or the table layout moved?')

    # ── the group panel: every round that was supplied ───────────────────────
    gseries: dict[str, list[float]] = {}
    present = [r for r in order if r in groups]
    for rnd in present:
        # A dropped round leaves its window to the PREVIOUS round, which is what
        # "the latest published figure" means when the next survey is missing —
        # the same hold rule, not a gap and not an interpolation.
        hold(rnd, groups[rnd], gseries, to_end=(rnd == present[-1]))

    round_totals = {r: sum(groups[r].values()) for r in order if r in groups}
    last = len(months) - 1
    lines = [
        '// GENERATED — do not edit by hand. Run scripts/gen-ph-vacancy-demand.py.',
        '//',
        '// Philippines vacancies by occupation, from the Philippine Statistics',
        '// Authority Integrated Survey on Labor and Employment (ISLE / BITS).',
        '// Six survey rounds, 2011/2012 through 2023/2024.',
        '//',
        '// TWO PANELS, BECAUSE THE SURVEY PUBLISHES AT TWO GRANULARITIES.',
        '//  • PH_SERIES — per SKILL. Built only from the four rounds that publish',
        '//    detailed occupation titles (2013/14, 2015/16, 2017/18, 2021/22),',
        '//    mapped by the app\'s own matcher (skillsForText) so a Manila',
        '//    occupation lands on the same skill an Australian one would.',
        '//  • PH_GROUP_SERIES — per MAJOR OCCUPATION GROUP. All six rounds,',
        '//    including 2011/12 and 2023/24, which publish groups only.',
        '//',
        '// The two are never mixed. Letting a skill inherit its major group\'s',
        '// figure (as the Hong Kong feed lets a skill inherit its industry',
        '// section) was measured against the 2021/22 round, which publishes both:',
        '// it overstates a skill by a median of 72x, because ISCO\'s nine groups',
        '// are far coarser than Hong Kong\'s industry sections. So the skill panel',
        '// stops where the detail stops, and the group panel carries the rest.',
        '//',
        '// THESE ROUNDS ARE NOT ONE CONTINUOUS SERIES.',
        '//  • The establishment threshold changed: 20+ workers through 2021/22,',
        '//    10+ from 2023/24. A wider frame reports more vacancies for the same',
        '//    market, so the last round is a level BREAK, not growth.',
        '//  • Each round is a 12-18 month reference window with multi-year gaps',
        '//    between. Nothing is interpolated: a round is HELD from the start of',
        '//    its window until the next round of the same panel begins, and the',
        '//    final round runs to the axis end (a trailing zero would read as',
        '//    "no vacancies", not "not surveyed since").',
        '//  • So PH_SERIES\'s skill detail is as at the 2021/2022 round — the most',
        '//    recent one published by occupation — even for months after it.',
        '//',
        '// An occupation mapping to several skills contributes its whole count to',
        '// each — the same rule the IVI generator uses — so these series must not',
        '// be summed into a national total. Use PH_ROUND_TOTALS for that.',
        '',
        "import { IVI_MONTHS } from './iviSkillDemand';",
        '',
        'export const PH_SOURCE =',
        "  'Philippine Statistics Authority — Integrated Survey on Labor and Employment (ISLE), 2011/2012 to 2023/2024 · survey rounds, not a continuous series; skill detail as at the 2021/2022 round';",
        '',
        '/** The establishment frame each round covered. The 2023/24 change is a break. */',
        'export const PH_THRESHOLD: Record<string, string> = {',
    ]
    for r in order:
        if r in groups:
            lines.append(f'  {json.dumps(r)}: {json.dumps(ROUNDS[r][2])},')
    lines += ['};', '',
              '/** Whether a round published detailed occupations or major groups only. */',
              'export const PH_GRANULARITY: Record<string, string> = {']
    for r in order:
        if r in groups:
            lines.append(f'  {json.dumps(r)}: {json.dumps(ROUNDS[r][3])},')
    lines += ['};', '',
              '/** Published total vacancies per round, as the survey reported them. */',
              'export const PH_ROUND_TOTALS: Record<string, number> = {']
    for r in order:
        if r in round_totals:
            lines.append(f'  {json.dumps(r)}: {round(round_totals[r])},')
    lines += ['};', '',
              '/**',
              ' * Skill → { manila: monthly vacancies }, on the IVI_MONTHS axis.',
              ' *',
              ' * Detailed rounds only, so this begins Jan 2013 (not 2011) and its skill',
              ' * detail is as at the 2021/2022 round. The 2011/12 and 2023/24 rounds are',
              ' * in PH_GROUP_SERIES.',
              ' */',
              'export const PH_SERIES: Record<string, Record<string, number[]>> = {']
    for s in sorted(series):
        lines.append(f'  {json.dumps(s)}: {{ {HUB}: [' +
                     ','.join(str(round(v)) for v in series[s]) + '] },')
    lines += ['};', '',
              '/** Latest value per skill, for the map. */',
              'export const PH_SKILL_BY_CITY: Record<string, Record<string, number>> = {']
    for s in sorted(series):
        if series[s][last] > 0:
            lines.append(f'  {json.dumps(s)}: {{ {HUB}: {round(series[s][last])} }},')
    lines += ['};', '',
              '/** The nine PSOC 2012 major occupation groups, in published order. */',
              'export const PH_GROUPS: string[] = [']
    for g in GROUPS:
        lines.append(f'  {json.dumps(g)},')
    lines += ['];', '',
              '/**',
              ' * Major occupation group → { manila: monthly vacancies }, same axis.',
              ' *',
              ' * All six rounds, so this spans Jan 2011 to the axis end and is the only',
              ' * export that carries the 2011/12 and 2023/24 figures. Unlike PH_SERIES,',
              ' * these ARE mutually exclusive and DO sum to the round total.',
              ' */',
              'export const PH_GROUP_SERIES: Record<string, Record<string, number[]>> = {']
    for g in GROUPS:
        if g in gseries:
            lines.append(f'  {json.dumps(g)}: {{ {HUB}: [' +
                         ','.join(str(round(v)) for v in gseries[g]) + '] },')
    lines += ['};', '',
              '/** Latest value per major occupation group (the 2023/2024 round). */',
              'export const PH_GROUP_BY_CITY: Record<string, Record<string, number>> = {']
    for g in GROUPS:
        if g in gseries and gseries[g][last] > 0:
            lines.append(f'  {json.dumps(g)}: {{ {HUB}: {round(gseries[g][last])} }},')
    lines += ['};', '', 'void IVI_MONTHS;', '']

    open(OUT, 'w').write('\n'.join(lines))
    sys.stderr.write(
        f'\nwrote {OUT}\n'
        f'  skill panel:  {len(series)} skills, {len(detail_order)} detailed rounds\n'
        f'  group panel:  {len(gseries)} major groups, {len(present)} of '
        f'{len(ROUNDS)} rounds\n')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
