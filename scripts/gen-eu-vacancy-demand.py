#!/usr/bin/env python3
"""Regenerate src/employsi/data/euVacancyDemand.ts from the Eurostat "Job vacancy
statistics by occupation and NUTS 1 region (OJA breakdowns) — annual, experimental"
release (jvs_a_isco3_r1), broken down by occupation (ISCO-08 3-digit) × geo × year.

Like the UK ONS wiring, each occupation label is term-matched to canonical skills
via the shared skillsTaxonomy terms and summed per skill. Unlike the UK (one city)
the user wants this recorded BY COUNTRY, so we keep the COUNTRY-level GEO rows
(dropping the EU/euro-area aggregates and the NUTS-1 sub-regions) and place each
country on its capital.

Converting to vacancies: the source figures are EMPLOYMENT by occupation (in
thousands — EU-27 Total ≈ 165M employees, Germany ≈ 37M), i.e. a stock, whereas
the UK ONS series (and the app's other feeds) are vacancy COUNTS. The conversion
uses Eurostat's OWN job-vacancy-rate series, fetched live from the dissemination
API (jvs_q_nace2, indicator JVR, NACE B-S "industry, construction and services",
unadjusted, all size classes) — the headline rate each member state reports,
quarterly, per country:

    vacancies(skill, country, month)
        = employment(skill, country, year-of-month) × r / (1 − r)

where r is that country's job vacancy rate for the quarter the month falls in.
Eurostat defines the job vacancy rate as vacancies ÷ (occupied posts +
vacancies), so r/(1 − r) is the inversion of that identity, not an adjustment
factor — the same arithmetic the US wiring uses with the JOLTS openings rate.

This replaces an earlier flat 2.5% applied to every country in every year, which
was a reasonable placeholder but wrong in both directions at once: the Czech
Republic and the Netherlands have run 3-5% while Spain, Greece and Romania sit
near 0.5-1%, so a single constant overstated southern Europe by several times
and understated the tight north-western markets.

The occupation release is annual (2019..latest) and the rate is quarterly, so we
emit the compact per-year employment and per-country quarterly rates, and let the
TS combine them onto the shared IVI_MONTHS axis at load time (employment held
flat across its calendar year and carried forward to "now"; the rate stepping
each quarter; pre-first-year months zero). This keeps the data file small while
conforming to the seriesFor contract every other country series uses.

Usage: python3 scripts/gen-eu-vacancy-demand.py path/to/jvs_a_isco3_r1.xlsx
Network required for the vacancy-rate series (no API key needed).
"""
import json, re, sys, urllib.parse, urllib.request, warnings
warnings.filterwarnings('ignore')
import openpyxl

sys.path.insert(0, __file__.rsplit('/', 1)[0])
from skills_taxonomy import load_skills, matcher  # noqa: E402

ROOT = __file__.rsplit('/scripts/', 1)[0]
TAX = f'{ROOT}/src/employsi/data/skillsTaxonomy.ts'
IVI = f'{ROOT}/src/employsi/data/iviSkillDemand.ts'
OUT = f'{ROOT}/src/employsi/data/euVacancyDemand.ts'
SHEET = 'Data'

# EU country → canonical map id + display label + capital [lng, lat] + Eurostat
# geo code (used to pull that country's job vacancy rate). Only these GEO rows are
# kept (aggregates like "European Union - 27 countries" / "Euro area" and the
# NUTS-1 sub-regions / "Not regionalised" are dropped). Denmark carries no rows in
# this extract, so it's absent. Keyed by the exact Eurostat GEO label.
# Note Greece is 'EL' in Eurostat's code list, not the ISO 'GR'.
COUNTRIES = {
    'Belgium':     ('belgium',     'Belgium',        [4.35, 50.85], 'BE'),
    'Bulgaria':    ('bulgaria',    'Bulgaria',       [23.32, 42.70], 'BG'),
    'Czechia':     ('czechia',     'Czechia',        [14.42, 50.08], 'CZ'),
    'Germany':     ('germany',     'Germany',        [13.40, 52.52], 'DE'),
    'Estonia':     ('estonia',     'Estonia',        [24.75, 59.44], 'EE'),
    'Ireland':     ('ireland',     'Ireland',        [-6.26, 53.35], 'IE'),
    'Greece':      ('greece',      'Greece',         [23.73, 37.98], 'EL'),
    'Spain':       ('spain',       'Spain',          [-3.70, 40.42], 'ES'),
    'France':      ('france',      'France',         [2.35, 48.86], 'FR'),
    'Croatia':     ('croatia',     'Croatia',        [15.98, 45.81], 'HR'),
    'Italy':       ('italy',       'Italy',          [12.50, 41.90], 'IT'),
    'Cyprus':      ('cyprus',      'Cyprus',         [33.36, 35.17], 'CY'),
    'Latvia':      ('latvia',      'Latvia',         [24.11, 56.95], 'LV'),
    'Lithuania':   ('lithuania',   'Lithuania',      [25.28, 54.69], 'LT'),
    'Luxembourg':  ('luxembourg',  'Luxembourg',     [6.13, 49.61], 'LU'),
    'Hungary':     ('hungary',     'Hungary',        [19.04, 47.50], 'HU'),
    'Malta':       ('malta',       'Malta',          [14.51, 35.90], 'MT'),
    'Netherlands': ('netherlands', 'Netherlands',    [4.90, 52.37], 'NL'),
    'Austria':     ('austria',     'Austria',        [16.37, 48.21], 'AT'),
    'Poland':      ('poland',      'Poland',         [21.01, 52.23], 'PL'),
    'Portugal':    ('portugal',    'Portugal',       [-9.14, 38.72], 'PT'),
    'Romania':     ('romania',     'Romania',        [26.10, 44.43], 'RO'),
    'Slovenia':    ('slovenia',    'Slovenia',       [14.51, 46.06], 'SI'),
    'Slovakia':    ('slovakia',    'Slovakia',       [17.11, 48.15], 'SK'),
    'Finland':     ('finland',     'Finland',        [24.94, 60.17], 'FI'),
    'Sweden':      ('sweden',      'Sweden',         [18.07, 59.33], 'SE'),
}

# Occupation rows to ignore: the "Total" line and the ICT custom aggregate (it
# double-counts occupations already itemised).
SKIP_OCC = {'total'}

# Eurostat's own job-vacancy-rate series, used to turn the employment stock into
# vacancy counts. NACE B-S ("industry, construction and services") is the headline
# aggregate every member state publishes — A-S is missing for eight of the 26,
# including France, Spain, Italy and Austria, so B-S is what makes the set
# complete and comparable.
JVR_DATASET = 'https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/jvs_q_nace2'
JVR_QUERY = {
    'format': 'JSON', 'lang': 'en',
    'indic_em': 'JVR',        # job vacancy rate (vacancies / (occupied + vacancies))
    'nace_r2': 'B-S',         # industry, construction and services
    's_adj': 'NSA',           # unadjusted — the published headline
    'sizeclas': 'TOTAL',      # all size classes
}


def load_ivi_months():
    """The shared IVI_MONTHS axis every country series is aligned to."""
    m = re.search(r'IVI_MONTHS[^=]*=\s*(\[.*?\]);', open(IVI).read(), re.S)
    if not m:
        sys.exit('Could not read IVI_MONTHS from iviSkillDemand.ts')
    return json.loads(m.group(1).replace("'", '"'))


def fetch_vacancy_rates(geos):
    """{eurostat_geo: {'YYYY-Qn': rate_pct}} straight from the dissemination API."""
    url = (JVR_DATASET + '?' + urllib.parse.urlencode(JVR_QUERY)
           + ''.join(f'&geo={g}' for g in sorted(geos)))
    d = json.loads(urllib.request.urlopen(url, timeout=120).read())
    if 'error' in d:
        sys.exit(f'Eurostat refused the vacancy-rate request: {d["error"]}')
    gidx = d['dimension']['geo']['category']['index']
    tidx = d['dimension']['time']['category']['index']
    inv_g = {v: k for k, v in gidx.items()}
    inv_t = {v: k for k, v in tidx.items()}
    nt = len(tidx)
    out = {}
    # jsonstat flattens [geo][time] into one index; only the two dimensions above
    # have more than one category, so the stride is simply the number of periods.
    for k, v in d['value'].items():
        k = int(k)
        out.setdefault(inv_g[k // nt], {})[inv_t[k % nt]] = float(v)
    return out


def quarter_of(ym):
    """'2024-07' -> '2024-Q3' (Eurostat's period label)."""
    return f'{ym[:4]}-Q{(int(ym[5:7]) - 1) // 3 + 1}'


def main(path):
    skills_for = matcher(TAX)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET]
    nrow, ncol = ws.max_row, ws.max_column

    # Locate the TIME row (col A == 'TIME'), the GEO labels row, and the first
    # occupation row (the one after the 'ISCO08 (Labels)' header).
    time_row = geo_row = isco_row = None
    for i in range(1, nrow + 1):
        a = str(ws.cell(i, 1).value or '').strip()
        if a == 'TIME':
            time_row = i
        elif a.startswith('GEO'):
            geo_row = i
        elif a.startswith('ISCO'):
            isco_row = i
            break
    if not (time_row and geo_row and isco_row):
        sys.exit(f'Could not find TIME/GEO/ISCO header rows ({time_row},{geo_row},{isco_row}).')

    # TIME markers: {col: year}. Each year block spans until the next marker.
    marks = [(j, int(ws.cell(time_row, j).value)) for j in range(2, ncol + 1)
             if str(ws.cell(time_row, j).value or '').strip().isdigit()]
    years = [y for _, y in marks]
    blocks = []  # (year, start_col, end_col)
    for k, (col, yr) in enumerate(marks):
        end = marks[k + 1][0] - 1 if k + 1 < len(marks) else ncol
        blocks.append((yr, col, end))

    # Within any block the GEO columns repeat in the same order; grab the country
    # value-columns (label matches a COUNTRIES key) relative to the block start.
    first_start = blocks[0][1]
    country_cols = []  # (rel_offset, country_key)
    for j in range(first_start, blocks[0][2] + 1):
        lbl = str(ws.cell(geo_row, j).value or '').strip()
        if lbl in COUNTRIES:
            country_cols.append((j - first_start, lbl))

    # acc[skill][country_key][year] = summed count
    acc = {}
    for i in range(isco_row + 1, nrow + 1):
        occ = str(ws.cell(i, 1).value or '').strip()
        if not occ or occ.lower() in SKIP_OCC:
            continue
        if occ.lower().startswith('information and communications technology specialists'):
            continue  # custom ICT aggregate — would double-count
        sk = skills_for(occ)
        if not sk:
            continue
        for yr, start, _end in blocks:
            for off, ckey in country_cols:
                cell = ws.cell(i, start + off).value
                try:
                    v = float(cell)
                except (TypeError, ValueError):
                    continue  # ':' confidential / missing
                # thousands of employees → employees. The conversion to
                # vacancies happens in the TS, per country per quarter, using
                # that country's real Eurostat job vacancy rate.
                cnt = round(v * 1000)
                if cnt <= 0:
                    continue
                for s in sk:
                    acc.setdefault(s, {}).setdefault(ckey, {}).setdefault(yr, 0)
                    acc[s][ckey][yr] += cnt

    # ── Eurostat's own job vacancy rate, per country per quarter ─────────────
    sys.stderr.write('Fetching Eurostat job vacancy rates (jvs_q_nace2 / JVR / B-S)…\n')
    geo_of = {k: COUNTRIES[k][3] for k in COUNTRIES}
    rates = fetch_vacancy_rates(set(geo_of.values()))
    missing = sorted(k for k in COUNTRIES if not rates.get(geo_of[k]))
    if missing:
        sys.exit(f'No job vacancy rate published for {missing} — refusing to fall '
                 f'back to a flat rate silently.')
    # The occupation release starts at years[0]; anything earlier is unused.
    months = load_ivi_months()
    quarters = sorted({quarter_of(m) for m in months if int(m[:4]) >= years[0]})

    latest_year = max(years)

    def latest_rate(ckey):
        """Most recent published rate for a country, as a fraction."""
        by_q = rates[geo_of[ckey]]
        return by_q[max(by_q)] / 100

    # Order skills by latest-year VACANCIES (not employment) across all countries,
    # so the emitted order matches what the app actually ranks by.
    def skill_total(s):
        t = 0
        for ckey, cy in acc[s].items():
            e = cy.get(latest_year, 0)
            r = latest_rate(ckey)
            if e > 0 and 0 < r < 1:
                t += round(e * r / (1 - r))
        return t
    order = sorted(acc, key=lambda s: -skill_total(s))

    id_of = {k: COUNTRIES[k][0] for k in COUNTRIES}

    L = []
    A = L.append
    A('// GENERATED — do not edit by hand. Run scripts/gen-eu-vacancy-demand.py.')
    A('// Sources: Eurostat —')
    A('//   • Job vacancy statistics by occupation and NUTS 1 region (OJA breakdowns),')
    A('//     annual experimental statistics [jvs_a_isco3_r1]: EMPLOYMENT by ISCO-08')
    A('//     occupation × country × year. Each occupation is term-matched to canonical')
    A('//     skills (shared skillsTaxonomy terms) and summed per skill, BY COUNTRY')
    A('//     (aggregates + NUTS-1 sub-regions dropped).')
    A('//   • Job vacancy statistics by NACE Rev. 2 activity, quarterly [jvs_q_nace2],')
    A('//     indicator JVR, NACE B-S, unadjusted: each country\'s own published JOB')
    A('//     VACANCY RATE, per quarter.')
    A('//')
    A('// The occupation release is an employment stock, so on its own it is not')
    A("// comparable with the UK/AU/US feeds, which are vacancy counts. Eurostat")
    A('// defines the job vacancy rate r as vacancies ÷ (occupied posts + vacancies),')
    A('// so vacancies are recovered from an employment base as employment × r / (1 − r)')
    A('// — the inversion of that identity, not an adjustment factor. Each country uses')
    A("// its OWN rate for the quarter in question (the US wiring does the same with")
    A('// the JOLTS state openings rate).')
    A('//')
    A('// This replaced a flat 2.5% applied to every country in every year. That single')
    A('// constant was wrong in both directions at once — Czechia and the Netherlands')
    A('// run several times higher than Spain, Greece or Romania — so it overstated')
    A('// southern Europe and understated the tight north-western markets.')
    A('//')
    A('// Annual employment is held flat across its calendar year and carried forward to')
    A('// "now"; the rate steps each quarter; months before the first year are zero.')
    A('')
    A("import { IVI_MONTHS } from './iviSkillDemand';")
    A('')
    A(f'export const EU_YEARS: number[] = {json.dumps(years)};')
    A("export const EU_SOURCE = 'Eurostat — employment by occupation (OJA breakdowns) "
      "scaled to vacancies by each country\\'s own job vacancy rate';")
    A('')
    A('// Country id → display label and capital [lng, lat] (the marker + heat location).')
    A('export const EU_CITY_LABEL: Record<string, string> = {')
    for k in COUNTRIES:
        A(f'  {COUNTRIES[k][0]}: {json.dumps(COUNTRIES[k][1])},')
    A('};')
    A('export const EU_CITY_LNGLAT: Record<string, [number, number]> = {')
    for k in COUNTRIES:
        lng, lat = COUNTRIES[k][2]
        A(f'  {COUNTRIES[k][0]}: [{lng}, {lat}],')
    A('};')
    A(f'export const EU_CITIES: string[] = {json.dumps([COUNTRIES[k][0] for k in COUNTRIES])};')
    A('')
    A('// Quarter labels covering the axis from EU_YEARS[0] onward.')
    A(f'const EU_QUARTERS: string[] = {json.dumps(quarters)};')
    A('')
    A("// Country id → job vacancy rate (%) per EU_QUARTERS. 0 marks a quarter the")
    A('// country did not publish; expand() forward-fills those from the previous one.')
    A('const EU_VACANCY_RATE: Record<string, number[]> = {')
    for k in COUNTRIES:
        by_q = rates[geo_of[k]]
        arr = [by_q.get(q, 0) for q in quarters]
        A(f'  {COUNTRIES[k][0]}: [{",".join(f"{v:g}" for v in arr)}],')
    A('};')
    A('')
    A('// Compact source values: skill → country id → [EMPLOYMENT per EU_YEARS].')
    A('const EU_EMPLOYMENT: Record<string, Record<string, number[]>> = {')
    for s in order:
        parts = []
        for k in COUNTRIES:
            if k in acc[s]:
                arr = [acc[s][k].get(y, 0) for y in years]
                if any(arr):
                    parts.append(f'{id_of[k]}: [{",".join(map(str, arr))}]')
        if parts:
            A(f'  {json.dumps(s)}: {{ {", ".join(parts)} }},')
    A('};')
    A('')
    A("// Month index → index into EU_YEARS (most recent vintage at or before that")
    A('// month, -1 before the first) and into EU_QUARTERS. Computed once for the axis.')
    A('const YEAR_SLOT: number[] = IVI_MONTHS.map((ym) => {')
    A('  const y = Number(ym.slice(0, 4));')
    A('  let slot = -1;')
    A('  for (let i = 0; i < EU_YEARS.length; i++) if (EU_YEARS[i] <= y) slot = i;')
    A('  return slot;')
    A('});')
    A('const QUARTER_SLOT: number[] = IVI_MONTHS.map((ym) => {')
    A('  const q = `${ym.slice(0, 4)}-Q${Math.floor((Number(ym.slice(5, 7)) - 1) / 3) + 1}`;')
    A('  const i = EU_QUARTERS.indexOf(q);')
    A('  // Months past the last published quarter carry the final one forward.')
    A('  return i >= 0 ? i : q > (EU_QUARTERS[EU_QUARTERS.length - 1] || "") ? EU_QUARTERS.length - 1 : -1;')
    A('});')
    A('')
    A('// Forward-filled rate per country: a quarter a country did not publish inherits')
    A('// the most recent one it did, so a reporting gap holds the level rather than')
    A('// dropping the country off the map for a quarter.')
    A('const RATE_FILLED: Record<string, number[]> = Object.fromEntries(')
    A('  Object.entries(EU_VACANCY_RATE).map(([c, arr]) => {')
    A('    let last = 0;')
    A('    return [c, arr.map((v) => (v > 0 ? (last = v) : last))];')
    A('  }),')
    A(');')
    A('')
    A('// Employment → active vacancies, per country per month, at that country\'s own rate.')
    A('function expand(byCity: Record<string, number[]>): Record<string, number[]> {')
    A('  const out: Record<string, number[]> = {};')
    A('  for (const [city, emp] of Object.entries(byCity)) {')
    A('    const rate = RATE_FILLED[city] || [];')
    A('    out[city] = IVI_MONTHS.map((_ym, i) => {')
    A('      const yi = YEAR_SLOT[i];')
    A('      const qi = QUARTER_SLOT[i];')
    A('      if (yi < 0 || qi < 0) return 0;')
    A('      const e = emp[yi] || 0;')
    A('      const r = (rate[qi] || 0) / 100;')
    A('      if (e <= 0 || r <= 0 || r >= 1) return 0;')
    A('      return Math.round((e * r) / (1 - r));')
    A('    });')
    A('  }')
    A('  return out;')
    A('}')
    A('')
    A('// Skill → country id → monthly vacancy history (aligned to IVI_MONTHS).')
    A('export const EU_SERIES: Record<string, Record<string, number[]>> =')
    A('  Object.fromEntries(Object.entries(EU_EMPLOYMENT).map(([s, e]) => [s, expand(e)]));')
    A('')
    A('// Skill → latest-month country vacancy count (current heat map).')
    A('export const EU_SKILL_BY_CITY: Record<string, Record<string, number>> =')
    A('  Object.fromEntries(')
    A('    Object.entries(EU_SERIES).map(([s, byCity]) => [')
    A('      s,')
    A('      Object.fromEntries(Object.entries(byCity).map(([c, arr]) => [c, arr[arr.length - 1]])),')
    A('    ]),')
    A('  );')
    A('')
    open(OUT, 'w').write('\n'.join(L))

    ncountries = len({c for s in acc for c in acc[s]})
    tot = sum(skill_total(s) for s in order)
    print(f'{len(order)} skills mapped across {ncountries} countries, years {years[0]}..{years[-1]}, '
          f'latest-quarter vacancies {tot:,} -> {OUT}')
    lo = sorted(((latest_rate(k) * 100, COUNTRIES[k][1]) for k in COUNTRIES))
    print('latest published vacancy rates — lowest: '
          + ', '.join(f'{n} {r:.1f}%' for r, n in lo[:3])
          + ' | highest: ' + ', '.join(f'{n} {r:.1f}%' for r, n in lo[-3:]))


if __name__ == '__main__':
    if len(sys.argv) != 2:
        sys.exit('usage: gen-eu-vacancy-demand.py path/to/jvs_a_isco3_r1.xlsx')
    main(sys.argv[1])
